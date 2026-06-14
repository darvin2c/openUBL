"""Auditoría de cobertura de reglas SUNAT ERROR localmente evaluables.

Lee el Excel oficial de SUNAT, los validadores implementados y los tests,
y emite coverage_report.json con el estado por tipo de documento.
"""

from __future__ import annotations

import ast
import json
import re
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent

EXCEL_URL = (
    "https://cpe.sunat.gob.pe/sites/default/files/2026-05/"
    "Reglas%20de%20validaci%C3%B3n%20actualizado%20al%2024.04.2026%20%281%29.xlsx"
)

SHEET_TO_DOC: dict[str, str] = {
    "Factura2_0": "Invoice",
    "Boleta2_0": "Invoice",
    "NotaCredito2_0": "CreditNote",
    "NotaDebito2_0": "DebitNote",
    "Comunicación de Baja1_0": "VoidedDocuments",
    "Resumen Diario1_1": "SummaryDocuments",
    "Percepciones1_0": "Perception",
    "Retenciones1_0": "Retention",
    "Firma": "Signature",
}

GENERAL_CODES: set[str] = {
    "0100", "0109", "0111", "0151", "0154", "0155", "0156", "0157", "0158", "0160", "0161", "0306"
}

# Set inicial de códigos fuera de alcance. Puede ampliarse durante la implementación
# si una regla resulta depender de listados/padrones no disponibles localmente.
FUERA_DE_ALCANCE: dict[str, set[str]] = {
    "Invoice": {
        "1032", "1033", "1034", "1035", "1036", "1078", "1079", "1080", "1083", "1084", "1086",
        "2010", "2011", "2036", "2040", "2041", "2108", "2329", "2377", "2505", "2520", "2529",
        "2788", "2792", "2798", "2800", "2961", "2964", "3007", "3027", "3033", "3071", "3097",
        "3116", "3134", "3150", "3174", "3181", "3207", "3218", "3219", "3239", "3240", "3269",
        "3281", "3283", "3284", "3285", "3289", "3496",
    },
    "CreditNote": {
        "1032", "1033", "1034", "1035", "1036", "1078", "1079", "1080", "1083", "1084", "1086",
        "2010", "2011", "2016", "2036", "2040", "2108", "2119", "2120", "2121", "2172", "2199",
        "2329", "2377", "2885", "2961", "2964", "3007", "3027", "3150", "3207", "3209", "3239",
        "3260", "3286", "3496", "3503",
    },
    "DebitNote": {
        "1032", "1033", "1034", "1035", "1036", "1078", "1079", "1080", "1083", "1084", "1086",
        "2010", "2011", "2016", "2036", "2040", "2108", "2172", "2199", "2207", "2208", "2209",
        "2329", "2377", "2885", "2961", "2964", "3007", "3027", "3033", "3150", "3174", "3207",
        "3209", "3239", "3496",
    },
    "VoidedDocuments": {"2011", "2324", "4203"},
    "SummaryDocuments": {
        "1078", "2016", "2223", "2256", "2268", "2282", "2517", "2601", "2663", "2891", "2896",
        "2987", "3207",
    },
    "Perception": {
        "1033", "1034", "1049", "2600", "2602", "2603", "2605", "2609", "2610", "3207", "3312",
        "3325", "3326", "3328", "3329",
    },
    "Retention": {"1033", "1034", "1049", "2600", "2617", "2618", "2619", "2621", "3207"},
    "Signature": set(),
}


def _normalize_header(cell) -> str:
    if cell is None:
        return ""
    text = str(cell).strip().strip("'\"")
    return text.upper().replace(" ", "").replace("\n", "")


def _find_columns(header_row):
    tipo_idx = None
    codigo_idx = None
    desc_idx = None
    for idx, cell in enumerate(header_row):
        norm = _normalize_header(cell)
        if norm in {"TIPODERETORNO", "TIPODERETORNOOBSERV/ERROR"}:
            tipo_idx = idx
        if norm in {"CODIGORETORNO", "CODIGODERETORNO", "CODIGO"}:
            codigo_idx = idx
        if norm in {"MENSAJEDERETORNO", "DESCRIPCION", "DESCRIPCIÓN", "DESCRIPCIONDECODIGODERETORNO"}:
            desc_idx = idx
    return tipo_idx, codigo_idx, desc_idx


def _download_excel() -> Path:
    cache_dir = REPO_ROOT / ".cache"
    cache_dir.mkdir(exist_ok=True)
    cache_path = cache_dir / "sunat_reglas_2026.xlsx"
    if cache_path.exists():
        return cache_path
    print(f"Descargando {EXCEL_URL} ...", file=sys.stderr)
    req = urllib.request.Request(EXCEL_URL, headers={"User-Agent": "openubl-coverage/1.0"})
    with urllib.request.urlopen(req, timeout=60) as response, open(cache_path, "wb") as f:
        f.write(response.read())
    return cache_path


def _load_sunat_codes(excel_path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    by_doc: dict[str, dict[str, str]] = {}
    for sheet_name in wb.sheetnames:
        doc = SHEET_TO_DOC.get(sheet_name.strip())
        if doc is None:
            continue
        ws = wb[sheet_name]
        header_indices = None
        for row in ws.iter_rows(values_only=True):
            if header_indices is None:
                tipo_idx, codigo_idx, desc_idx = _find_columns(row)
                if tipo_idx is not None and codigo_idx is not None:
                    header_indices = (tipo_idx, codigo_idx, desc_idx)
                continue
            tipo_idx, codigo_idx, desc_idx = header_indices
            tipo_val = row[tipo_idx]
            codigo_val = row[codigo_idx]
            desc_val = row[desc_idx] if desc_idx is not None else ""
            tipo = str(tipo_val).strip().strip("'\"").upper() if tipo_val is not None else ""
            codigo = str(codigo_val).strip().strip("'\"")
            desc = str(desc_val).strip() if desc_val is not None else ""
            if not codigo or not codigo.isdigit():
                continue
            if "ERROR" not in tipo:
                continue
            by_doc.setdefault(doc, {})[codigo] = desc
    wb.close()
    return by_doc


def _literal_codes_from_ast(path: Path) -> set[str]:
    if not path.exists():
        return set()
    text = path.read_text(encoding="utf-8")
    try:
        tree = ast.parse(text)
    except SyntaxError:
        return set()
    codes: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Call):
            for arg in node.args + [kw.value for kw in node.keywords]:
                if isinstance(arg, ast.Constant) and isinstance(arg.value, str):
                    val = arg.value.strip()
                    if val.isdigit() and len(val) == 4:
                        codes.add(val)
    return codes


def _codes_by_method(path: Path) -> dict[str, set[str]]:
    """Extrae códigos SUNAT agrupados por función/método en un archivo Python."""
    result: dict[str, set[str]] = {}
    if not path.exists():
        return result
    text = path.read_text(encoding="utf-8")
    try:
        tree = ast.parse(text)
    except SyntaxError:
        return result
    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef):
            codes: set[str] = set()
            for child in ast.walk(node):
                if isinstance(child, ast.Call):
                    for arg in child.args + [kw.value for kw in child.keywords]:
                        if isinstance(arg, ast.Constant) and isinstance(arg.value, str):
                            val = arg.value.strip()
                            if val.isdigit() and len(val) == 4:
                                codes.add(val)
            if codes:
                result[node.name] = codes
    return result


def _implemented_codes() -> dict[str, set[str]]:
    validators_dir = REPO_ROOT / "src" / "openubl" / "validators"
    validator_py = REPO_ROOT / "src" / "openubl" / "validator.py"

    method_codes = _codes_by_method(validator_py)

    common_methods = {
        "_validate_invoice_common",
        "_check_currency_consistency",
        "_check_tax_total",
        "_check_line_extension_amount",
    }
    common_codes: set[str] = set()
    for method in common_methods:
        common_codes |= method_codes.get(method, set())

    file_codes: dict[str, set[str]] = {
        "invoice1": _literal_codes_from_ast(validators_dir / "_extra_invoice1.py"),
        "invoice2": _literal_codes_from_ast(validators_dir / "_extra_invoice2.py"),
        "invoice3": _literal_codes_from_ast(validators_dir / "_extra_invoice3.py"),
        "invoice4": _literal_codes_from_ast(validators_dir / "_extra_invoice4.py"),
        "credit_note": _literal_codes_from_ast(validators_dir / "_extra_credit_note.py"),
        "credit_note2": _literal_codes_from_ast(validators_dir / "_extra_credit_note2.py"),
        "debit_note": _literal_codes_from_ast(validators_dir / "_extra_debit_note.py"),
        "debit_note2": _literal_codes_from_ast(validators_dir / "_extra_debit_note2.py"),
        "voided_summary": _literal_codes_from_ast(validators_dir / "_extra_voided_summary.py"),
        "perception_retention": _literal_codes_from_ast(validators_dir / "_extra_perception_retention.py"),
        "perception_retention2": _literal_codes_from_ast(validators_dir / "_extra_perception_retention2.py"),
    }

    doc_codes: dict[str, set[str]] = {doc: set() for doc in SHEET_TO_DOC.values()}
    doc_codes["Invoice"] = (
        file_codes["invoice1"] | file_codes["invoice2"] | file_codes["invoice3"] | file_codes["invoice4"]
        | common_codes | method_codes.get("_validate_invoice_specific", set())
    )
    doc_codes["CreditNote"] = (
        file_codes["credit_note"] | file_codes["credit_note2"]
        | common_codes | method_codes.get("_validate_credit_note_specific", set())
    )
    doc_codes["DebitNote"] = (
        file_codes["debit_note"] | file_codes["debit_note2"]
        | common_codes | method_codes.get("_validate_debit_note_specific", set())
    )
    doc_codes["VoidedDocuments"] = (
        file_codes["voided_summary"] | method_codes.get("_validate_voided_documents", set())
    )
    doc_codes["SummaryDocuments"] = (
        file_codes["voided_summary"] | method_codes.get("_validate_summary_documents", set())
    )
    doc_codes["Perception"] = (
        file_codes["perception_retention"] | file_codes["perception_retention2"]
        | method_codes.get("_validate_perception", set())
    )
    doc_codes["Retention"] = (
        file_codes["perception_retention"] | file_codes["perception_retention2"]
        | method_codes.get("_validate_retention", set())
    )
    doc_codes["Signature"] = (
        _literal_codes_from_ast(validators_dir / "_extra_signature.py")
        | method_codes.get("_validate_signature", set())
    )

    return doc_codes


def _tested_codes() -> dict[str, set[str]]:
    tests_dir = REPO_ROOT / "tests"
    code_pattern = re.compile(r'["\'](\d{4})["\']')
    doc_files: dict[str, list[Path]] = {
        "Invoice": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_invoice1_extra.py",
            tests_dir / "_test_validator_invoice2_extra.py",
            tests_dir / "_test_validator_invoice3_extra.py",
            tests_dir / "_test_validator_invoice4_extra.py",
        ],
        "CreditNote": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_credit_note_extra.py",
        ],
        "DebitNote": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_debit_note_extra.py",
        ],
        "VoidedDocuments": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_voided_summary_extra.py",
        ],
        "SummaryDocuments": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_voided_summary_extra.py",
        ],
        "Perception": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_perception_retention_extra.py",
        ],
        "Retention": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_perception_retention_extra.py",
        ],
        "Signature": [
            tests_dir / "test_validator.py",
            tests_dir / "_test_validator_signature_extra.py",
        ],
    }
    tested: dict[str, set[str]] = {}
    for doc, files in doc_files.items():
        codes: set[str] = set()
        for path in files:
            if not path.exists():
                continue
            text = path.read_text(encoding="utf-8")
            for match in code_pattern.finditer(text):
                codes.add(match.group(1))
        tested[doc] = codes
    return tested


def build_report() -> dict[str, Any]:
    excel_path = _download_excel()
    sunat_by_doc = _load_sunat_codes(excel_path)
    implemented = _implemented_codes()
    tested = _tested_codes()

    report: dict[str, Any] = {
        "source": str(EXCEL_URL),
        "documents": {},
        "general_codes": sorted(GENERAL_CODES),
    }

    all_docs = sorted(set(SHEET_TO_DOC.values()))
    for doc in all_docs:
        sunat_codes = set(sunat_by_doc.get(doc, {}))
        oos = FUERA_DE_ALCANCE.get(doc, set()) | GENERAL_CODES
        local_eval = sunat_codes - oos
        impl = implemented.get(doc, set())
        test = tested.get(doc, set())
        local_missing = local_eval - impl
        oos_missing = oos - impl
        report["documents"][doc] = {
            "total_sunat_error": len(sunat_codes),
            "implemented": sorted(impl),
            "tested": sorted(test & impl),
            "local_missing": sorted(local_missing),
            "oos_missing": sorted(oos_missing),
            "out_of_scope": sorted(oos),
            "implementable": sorted(local_eval),
            "descriptions": {k: v for k, v in sorted(sunat_by_doc.get(doc, {}).items())},
        }
    return report


def main() -> int:
    report = build_report()
    out_path = REPO_ROOT / "coverage_report.json"
    out_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Reporte escrito en {out_path}")
    for doc, data in report["documents"].items():
        total = data["total_sunat_error"]
        impl = len(data["implemented"])
        test = len(data["tested"])
        missing = len(data["local_missing"])
        implementable = len(data["implementable"])
        pct_valid = (impl / implementable * 100) if implementable else 0.0
        pct_test = (test / impl * 100) if impl else 0.0
        print(
            f"{doc:20s} total={total:4d} impl={impl:4d} tested={test:4d} "
            f"missing={missing:4d} pct_valid={pct_valid:5.1f}% pct_test={pct_test:5.1f}%"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
